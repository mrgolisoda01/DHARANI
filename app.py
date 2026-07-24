"""
============================================================
 Mr. Golisoda Training Portal — Python (Flask) Backend
============================================================
 Self-contained LMS backend. Uses:
   - Flask     -> web server / backend
   - SQLite    -> database (single file: golisoda.db)
   - Werkzeug  -> secure password hashing

 DELIVERY 1 UPGRADE (admin dashboard + employee management)
   - Dashboard stats (totals, completion rate, avg score)
   - Full employee list (name, emp_id, designation, phone, role, status)
   - Edit employee / change role / delete employee
   - Admin-driven password reset (temporary password)
   - Bulk add employees via CSV
   All EXISTING routes (signup, login, approve, scores) are unchanged.

 DEFAULT ADMIN LOGIN (change after first login)
   Employee ID:  ADMIN
   Password:     Golisoda@2026
============================================================
"""

import os
import io
import csv
import re
from datetime import datetime
from urllib.parse import urlparse, unquote
from flask import (Flask, request, session, redirect, url_for,
                   render_template, jsonify, g)
from werkzeug.security import generate_password_hash, check_password_hash

import psycopg2
import psycopg2.extras

# ---------------------------------------------------------------
#  Basic setup
# ---------------------------------------------------------------
app = Flask(__name__)
# Strong random secret key (set for production). Keep this private.
app.secret_key = os.environ.get("SECRET_KEY", "9fe42eef27a6bfbbd6764513ed4d5b10ec0e2b4e803984c917b3d89cd8960016")

# ---------------------------------------------------------------
#  DATABASE: Supabase (PostgreSQL)
#  The connection string can be provided via the DATABASE_URL env var
#  (recommended on Render). If not set, the fallback below is used.
#  NOTE: the password contains an '@', which is handled safely because
#  we parse the parts explicitly rather than relying on URL parsing.
# ---------------------------------------------------------------
DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()

# Fallback connection details (used if DATABASE_URL is not set).
# Password has an '@' in it — kept as a literal here, not URL-encoded.
_PG = {
    "host": os.environ.get("PGHOST", "aws-1-ap-south-1.pooler.supabase.com"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "postgres"),
    "user": os.environ.get("PGUSER", "postgres.mpcsvagagbhimjloumgx"),
    "password": os.environ.get("PGPASSWORD", "mrgolisoda@123"),
}


def _pg_params():
    """Return connection kwargs for psycopg2, from DATABASE_URL if given,
    otherwise from the _PG dict. Handles passwords containing '@'."""
    if DATABASE_URL:
        # Parse carefully. If the URL has an '@' in the password, urlparse
        # can misread it, so we split on the LAST '@' before the host.
        raw = DATABASE_URL
        if raw.startswith("postgresql://") or raw.startswith("postgres://"):
            raw = raw.split("://", 1)[1]
        # raw is now like  user:pass@host:port/dbname  (pass may contain @)
        creds, _, hostpart = raw.rpartition("@")
        user, _, password = creds.partition(":")
        hostport, _, dbname = hostpart.partition("/")
        host, _, port = hostport.partition(":")
        return {
            "host": host,
            "port": int(port or "6543"),
            "dbname": dbname or "postgres",
            "user": unquote(user),
            "password": unquote(password),
        }
    return dict(_PG)


DEFAULT_ADMIN_ID = "ADMIN"
DEFAULT_ADMIN_PW = "Golisoda@2026"

# Roles allowed in the system
VALID_ROLES = ("staff", "instructor", "admin")

# ---------------------------------------------------------------
#  Supabase Storage (for direct file uploads)
#  Files (PDF, Word, PPT, images) are uploaded to a public bucket
#  and the portal stores the resulting public URL as the module link.
# ---------------------------------------------------------------
SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://mpcsvagagbhimjloumgx.supabase.co").rstrip("/")
SUPABASE_BUCKET = os.environ.get("SUPABASE_BUCKET", "training-files")
# Secret key used server-side to upload files. Set via env var on Render.
# Reads SUPABASE_SERVICE_KEY first (Syed's env var name), then SUPABASE_SECRET, then a hardcoded fallback.
SUPABASE_SECRET = (os.environ.get("SUPABASE_SERVICE_KEY", "").strip()
                   or os.environ.get("SUPABASE_SECRET", "").strip()
                   or "sb_secret_eyD_5ueAuZ8N-BsyRIP6Ag_q42qjTxv")

# File types allowed for direct upload (everything the user asked for)
ALLOWED_UPLOAD_EXT = {
    ".pdf", ".doc", ".docx", ".ppt", ".pptx", ".xls", ".xlsx",
    ".png", ".jpg", ".jpeg", ".gif", ".webp", ".txt"
}
MAX_UPLOAD_MB = 45  # keep under the bucket's 50MB limit


# ---------------------------------------------------------------
#  SQLite-compatibility layer over PostgreSQL
#  The rest of this app was written for sqlite3 (using "?" placeholders
#  and db.execute(...).fetchone()). These wrappers let all that code keep
#  working unchanged on Postgres:
#    - convert "?" placeholders to "%s"
#    - make execute() return a cursor with fetchone/fetchall
#    - rows behave like dicts AND tuples (like sqlite3.Row)
# ---------------------------------------------------------------
def _q(sql):
    """Convert sqlite-style '?' placeholders to psycopg2 '%s'.
    Leaves already-'%s' strings alone and ignores '?' inside quotes."""
    out, in_s, in_d = [], False, False
    for ch in sql:
        if ch == "'" and not in_d:
            in_s = not in_s
        elif ch == '"' and not in_s:
            in_d = not in_d
        if ch == "?" and not in_s and not in_d:
            out.append("%s")
        else:
            out.append(ch)
    return "".join(out)


class _Cur:
    """Wraps a psycopg2 cursor to mimic sqlite3's execute().fetchone() usage."""
    def __init__(self, cur):
        self._c = cur

    def fetchone(self):
        return self._c.fetchone()

    def fetchall(self):
        return self._c.fetchall()

    @property
    def lastrowid(self):
        # emulate sqlite's lastrowid via RETURNING id when available
        try:
            row = self._c.fetchone()
            if row is not None:
                return row["id"] if "id" in row.keys() else row[0]
        except Exception:
            pass
        return None


class _Conn:
    """Wraps a psycopg2 connection to mimic the sqlite3 connection API
    used across this app (db.execute(...), db.commit())."""
    def __init__(self, conn):
        self._conn = conn

    def execute(self, sql, params=()):
        cur = self._conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(_q(sql), params)
        return _Cur(cur)

    def executescript(self, sql):
        cur = self._conn.cursor()
        cur.execute(sql)
        cur.close()

    def commit(self):
        self._conn.commit()

    def rollback(self):
        self._conn.rollback()

    def close(self):
        self._conn.close()

    @property
    def raw(self):
        return self._conn


def get_db():
    """Open a database connection for this request (Postgres/Supabase)."""
    if "db" not in g:
        conn = psycopg2.connect(connect_timeout=15, **_pg_params())
        conn.autocommit = False
        g.db = _Conn(conn)
    return g.db


@app.teardown_appcontext
def close_db(exception):
    db = g.pop("db", None)
    if db is not None:
        try:
            db.close()
        except Exception:
            pass


_db_ready = False
@app.before_request
def _ensure_db():
    global _db_ready
    if not _db_ready:
        init_db()
        _db_ready = True


def _column_exists(db, table, column):
    row = db.execute(
        "SELECT 1 FROM information_schema.columns WHERE table_name=? AND column_name=?",
        (table, column)
    ).fetchone()
    return row is not None


def init_db():
    """Create tables and the default admin if they do not exist (PostgreSQL).
    Safe to run every startup — uses CREATE TABLE IF NOT EXISTS."""
    conn = psycopg2.connect(connect_timeout=15, **_pg_params())
    conn.autocommit = True
    db = _Conn(conn)

    db.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id            SERIAL PRIMARY KEY,
            emp_id        TEXT UNIQUE NOT NULL,
            name          TEXT NOT NULL,
            phone         TEXT,
            designation   TEXT,
            password_hash TEXT NOT NULL,
            status        TEXT NOT NULL DEFAULT 'pending',
            role          TEXT NOT NULL DEFAULT 'staff',
            created_at    TEXT,
            must_reset    INTEGER NOT NULL DEFAULT 0,
            last_login    TEXT
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS scores (
            id         SERIAL PRIMARY KEY,
            emp_id     TEXT NOT NULL,
            module_id  TEXT NOT NULL,
            set_no     INTEGER NOT NULL,
            score      INTEGER NOT NULL,
            total      INTEGER NOT NULL,
            percent    INTEGER NOT NULL,
            passed     INTEGER NOT NULL,
            taken_at   TEXT
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS assessments (
            id            SERIAL PRIMARY KEY,
            title         TEXT NOT NULL,
            description   TEXT,
            roles         TEXT NOT NULL DEFAULT 'all',
            num_questions INTEGER NOT NULL DEFAULT 10,
            pass_percent  INTEGER NOT NULL DEFAULT 90,
            time_limit    INTEGER NOT NULL DEFAULT 0,
            active        INTEGER NOT NULL DEFAULT 1,
            created_by    TEXT,
            created_at    TEXT
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS questions (
            id            SERIAL PRIMARY KEY,
            assessment_id INTEGER NOT NULL,
            question      TEXT NOT NULL,
            opt_a         TEXT NOT NULL,
            opt_b         TEXT NOT NULL,
            opt_c         TEXT,
            opt_d         TEXT,
            correct       TEXT NOT NULL,
            category      TEXT
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS assessment_results (
            id            SERIAL PRIMARY KEY,
            assessment_id INTEGER NOT NULL,
            emp_id        TEXT NOT NULL,
            score         INTEGER NOT NULL,
            total         INTEGER NOT NULL,
            percent       INTEGER NOT NULL,
            passed        INTEGER NOT NULL,
            taken_at      TEXT,
            time_taken    INTEGER DEFAULT 0
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS answer_details (
            id            SERIAL PRIMARY KEY,
            result_id     INTEGER NOT NULL,
            assessment_id INTEGER NOT NULL,
            emp_id        TEXT NOT NULL,
            question_id   INTEGER,
            question_text TEXT,
            chosen        TEXT,
            correct       TEXT,
            is_correct    INTEGER NOT NULL DEFAULT 0,
            category      TEXT,
            taken_at      TEXT
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS content_modules (
            id           SERIAL PRIMARY KEY,
            kind         TEXT NOT NULL DEFAULT 'induction',
            title        TEXT NOT NULL,
            description  TEXT,
            link         TEXT NOT NULL,
            file_type    TEXT,
            min_minutes  INTEGER NOT NULL DEFAULT 0,
            roles        TEXT NOT NULL DEFAULT 'all',
            sort_order   INTEGER NOT NULL DEFAULT 0,
            status       TEXT NOT NULL DEFAULT 'live',
            created_by   TEXT,
            created_at   TEXT
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS videos (
            id           SERIAL PRIMARY KEY,
            title        TEXT NOT NULL,
            description  TEXT,
            link         TEXT NOT NULL,
            roles        TEXT NOT NULL DEFAULT 'all',
            sort_order   INTEGER NOT NULL DEFAULT 0,
            status       TEXT NOT NULL DEFAULT 'live',
            created_by   TEXT,
            created_at   TEXT
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS module_completions (
            id          SERIAL PRIMARY KEY,
            module_id   INTEGER NOT NULL,
            emp_id      TEXT NOT NULL,
            completed_at TEXT
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS video_completions (
            id          SERIAL PRIMARY KEY,
            video_id    INTEGER NOT NULL,
            emp_id      TEXT NOT NULL,
            completed_at TEXT
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS certificate_tracks (
            id            SERIAL PRIMARY KEY,
            cert_name     TEXT NOT NULL,
            kind          TEXT NOT NULL DEFAULT 'training',
            roles         TEXT NOT NULL DEFAULT 'all',
            require_modules INTEGER NOT NULL DEFAULT 1,
            require_assessment_id INTEGER,
            status        TEXT NOT NULL DEFAULT 'live',
            created_by    TEXT,
            created_at    TEXT
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS issued_certificates (
            id          SERIAL PRIMARY KEY,
            track_id    INTEGER NOT NULL,
            emp_id      TEXT NOT NULL,
            cert_name   TEXT NOT NULL,
            issued_at   TEXT
        )
    """)

    # --- Safe migration for older tables (add columns if missing) ---
    db.execute("""
        CREATE TABLE IF NOT EXISTS designations (
            id         SERIAL PRIMARY KEY,
            name       TEXT UNIQUE NOT NULL,
            sort_order INTEGER DEFAULT 0,
            created_at TEXT
        )
    """)
    # seed the defaults the first time only
    seeded = db.execute("SELECT COUNT(*) c FROM designations").fetchone()["c"]
    if not seeded:
        for i, r in enumerate(["BDE", "BDM", "State Head", "RSM", "NSM", "Corporate", "Back Office"]):
            db.execute(
                "INSERT INTO designations (name, sort_order, created_at) VALUES (?,?,?)",
                (r, i, datetime.utcnow().isoformat())
            )

    if not _column_exists(db, "users", "must_reset"):
        db.execute("ALTER TABLE users ADD COLUMN must_reset INTEGER NOT NULL DEFAULT 0")
    if not _column_exists(db, "users", "last_login"):
        db.execute("ALTER TABLE users ADD COLUMN last_login TEXT")
    # add time_taken to assessment_results if an older DB doesn't have it
    if not _column_exists(db, "assessment_results", "time_taken"):
        db.execute("ALTER TABLE assessment_results ADD COLUMN time_taken INTEGER DEFAULT 0")

    # Create default admin if not present
    cur = db.execute("SELECT 1 FROM users WHERE emp_id = ?", (DEFAULT_ADMIN_ID,))
    if cur.fetchone() is None:
        db.execute(
            "INSERT INTO users (emp_id,name,phone,designation,password_hash,status,role,created_at) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (DEFAULT_ADMIN_ID, "Administrator", "", "Management / Leadership",
             generate_password_hash(DEFAULT_ADMIN_PW), "approved", "admin",
             datetime.utcnow().isoformat())
        )
    conn.close()


# ---------------------------------------------------------------
#  Auth helpers
# ---------------------------------------------------------------
def current_user():
    if "emp_id" not in session:
        return None
    db = get_db()
    return db.execute("SELECT * FROM users WHERE emp_id = ?",
                      (session["emp_id"],)).fetchone()


def login_required(view):
    from functools import wraps
    @wraps(view)
    def wrapped(*a, **k):
        if current_user() is None:
            return redirect(url_for("login_page"))
        return view(*a, **k)
    return wrapped


def admin_required(view):
    from functools import wraps
    @wraps(view)
    def wrapped(*a, **k):
        u = current_user()
        if u is None or u["role"] != "admin":
            return redirect(url_for("login_page"))
        return view(*a, **k)
    return wrapped


# ---------------------------------------------------------------
#  Pages
# ---------------------------------------------------------------
@app.route("/")
def login_page():
    if current_user():
        return redirect(url_for("portal_page"))
    return render_template("login.html")


@app.route("/portal")
@login_required
def portal_page():
    return render_template("portal.html", user=current_user())


@app.route("/admin")
@admin_required
def admin_page():
    return render_template("admin.html", user=current_user())


@app.route("/assessment")
@login_required
def assessment_page():
    return render_template("assessment.html", user=current_user())


# ---------------------------------------------------------------
#  API: signup / login / logout  (UNCHANGED behaviour)
# ---------------------------------------------------------------
@app.route("/api/signup", methods=["POST"])
def api_signup():
    d = request.get_json(force=True)
    name = (d.get("name") or "").strip()
    emp_id = (d.get("emp_id") or "").strip()
    phone = (d.get("phone") or "").strip()
    desg = (d.get("designation") or "").strip()
    pw = d.get("password") or ""

    if not all([name, emp_id, phone, desg, pw]):
        return jsonify(ok=False, msg="Please fill in all fields."), 400
    if len(pw) < 6:
        return jsonify(ok=False, msg="Password must be at least 6 characters."), 400

    db = get_db()
    exists = db.execute("SELECT 1 FROM users WHERE emp_id = ?", (emp_id,)).fetchone()
    if exists:
        return jsonify(ok=False, msg="This Employee ID is already registered. Try signing in."), 400

    db.execute(
        "INSERT INTO users (emp_id,name,phone,designation,password_hash,status,role,created_at) "
        "VALUES (?,?,?,?,?,?,?,?)",
        (emp_id, name, phone, desg, generate_password_hash(pw),
         "pending", "staff", datetime.utcnow().isoformat())
    )
    db.commit()
    return jsonify(ok=True, msg="Request submitted! An admin will approve your account shortly.")


@app.route("/api/login", methods=["POST"])
def api_login():
    d = request.get_json(force=True)
    emp_id = (d.get("emp_id") or "").strip()
    pw = d.get("password") or ""
    if not emp_id or not pw:
        return jsonify(ok=False, msg="Enter your Employee ID and password."), 400

    db = get_db()
    u = db.execute("SELECT * FROM users WHERE emp_id = ?", (emp_id,)).fetchone()
    if u is None or not check_password_hash(u["password_hash"], pw):
        return jsonify(ok=False, msg="Employee ID or password is incorrect."), 401
    if u["status"] == "pending":
        return jsonify(ok=False, msg="Your account is awaiting admin approval."), 403

    session["emp_id"] = u["emp_id"]
    # record last login time
    db.execute("UPDATE users SET last_login = ? WHERE emp_id = ?",
               (datetime.utcnow().isoformat(), u["emp_id"]))
    db.commit()
    dest = url_for("admin_page") if u["role"] == "admin" else url_for("portal_page")
    return jsonify(ok=True, redirect=dest)


@app.route("/api/change-password", methods=["POST"])
@login_required
def api_change_password():
    """Any logged-in user can change their own password (needs current password)."""
    u = current_user()
    d = request.get_json(force=True)
    current_pw = d.get("current_password") or ""
    new_pw = d.get("new_password") or ""
    if not current_pw or not new_pw:
        return jsonify(ok=False, msg="Enter your current and new password."), 400
    if not check_password_hash(u["password_hash"], current_pw):
        return jsonify(ok=False, msg="Your current password is incorrect."), 401
    if len(new_pw) < 6:
        return jsonify(ok=False, msg="New password must be at least 6 characters."), 400
    db = get_db()
    db.execute("UPDATE users SET password_hash=?, must_reset=0 WHERE emp_id=?",
               (generate_password_hash(new_pw), u["emp_id"]))
    db.commit()
    return jsonify(ok=True, msg="Password changed successfully.")


@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify(ok=True)


# ---------------------------------------------------------------
#  API: quiz score saving + progress  (UNCHANGED)
# ---------------------------------------------------------------
@app.route("/api/save-score", methods=["POST"])
@login_required
def api_save_score():
    d = request.get_json(force=True)
    u = current_user()
    score = int(d.get("score", 0))
    total = int(d.get("total", 0))
    percent = round((score / total) * 100) if total else 0
    passed = 1 if percent >= 90 else 0
    db = get_db()
    db.execute(
        "INSERT INTO scores (emp_id,module_id,set_no,score,total,percent,passed,taken_at) "
        "VALUES (?,?,?,?,?,?,?,?)",
        (u["emp_id"], d.get("module_id", ""), int(d.get("set_no", 1)),
         score, total, percent, passed, datetime.utcnow().isoformat())
    )
    db.commit()
    return jsonify(ok=True, percent=percent, passed=bool(passed))


@app.route("/api/my-progress")
@login_required
def api_my_progress():
    u = current_user()
    db = get_db()
    rows = db.execute(
        "SELECT module_id, MAX(percent) AS best, MAX(passed) AS passed "
        "FROM scores WHERE emp_id = ? GROUP BY module_id", (u["emp_id"],)
    ).fetchall()
    return jsonify(ok=True, progress={r["module_id"]: {"best": r["best"], "passed": bool(r["passed"])} for r in rows})


# ---------------------------------------------------------------
#  API: admin approve / reject  (UNCHANGED)
# ---------------------------------------------------------------
@app.route("/api/approve", methods=["POST"])
@admin_required
def api_approve():
    d = request.get_json(force=True)
    emp_id = (d.get("emp_id") or "").strip()
    db = get_db()
    db.execute("UPDATE users SET status='approved' WHERE emp_id = ?", (emp_id,))
    db.commit()
    return jsonify(ok=True)


@app.route("/api/reject", methods=["POST"])
@admin_required
def api_reject():
    d = request.get_json(force=True)
    emp_id = (d.get("emp_id") or "").strip()
    db = get_db()
    db.execute("DELETE FROM users WHERE emp_id = ? AND role != 'admin'", (emp_id,))
    db.commit()
    return jsonify(ok=True)


# ===============================================================
#  NEW — DELIVERY 1: Admin dashboard + employee management
# ===============================================================

@app.route("/api/admin/dashboard")
@admin_required
def api_admin_dashboard():
    """Stats for the dashboard home + full employee list."""
    db = get_db()

    total = db.execute("SELECT COUNT(*) c FROM users WHERE role != 'admin'").fetchone()["c"]
    pending = db.execute("SELECT COUNT(*) c FROM users WHERE status='pending'").fetchone()["c"]
    approved = db.execute("SELECT COUNT(*) c FROM users WHERE status='approved' AND role != 'admin'").fetchone()["c"]

    # learners who have passed at least one assessment / total approved learners
    passed_any = db.execute(
        "SELECT COUNT(DISTINCT emp_id) c FROM assessment_results WHERE passed = 1"
    ).fetchone()["c"]
    completion = round((passed_any / approved) * 100) if approved else 0

    avg_row = db.execute("SELECT AVG(percent) a FROM assessment_results").fetchone()
    avg_score = round(avg_row["a"]) if avg_row["a"] is not None else 0

    # full employee list (include admins too, marked, so all accounts are visible)
    rows = db.execute(
        "SELECT emp_id,name,phone,designation,role,status,created_at,last_login "
        "FROM users ORDER BY (status='pending') DESC, name"
    ).fetchall()
    employees = [dict(r) for r in rows]

    u = current_user()
    me = {
        "emp_id": u["emp_id"],
        "name": u["name"],
        "last_login": u["last_login"] if "last_login" in u.keys() else None,
    }

    return jsonify(
        ok=True,
        stats={"total": total, "pending": pending,
               "approved": approved, "completion": completion, "avg": avg_score},
        employees=employees,
        me=me,
    )


@app.route("/api/admin/update-user", methods=["POST"])
@admin_required
def api_admin_update_user():
    """Edit an employee's name, phone, designation, or role."""
    d = request.get_json(force=True)
    emp_id = (d.get("emp_id") or "").strip()
    if not emp_id:
        return jsonify(ok=False, msg="Missing employee ID."), 400

    db = get_db()
    target = db.execute("SELECT * FROM users WHERE emp_id = ?", (emp_id,)).fetchone()
    if target is None:
        return jsonify(ok=False, msg="Employee not found."), 404

    name = (d.get("name") or target["name"]).strip()
    phone = (d.get("phone") or target["phone"] or "").strip()
    desg = (d.get("designation") or target["designation"] or "").strip()
    role = (d.get("role") or target["role"]).strip()
    if role not in VALID_ROLES:
        return jsonify(ok=False, msg="Invalid role."), 400

    # Safety: never allow the last admin to be demoted (lockout protection)
    if target["role"] == "admin" and role != "admin":
        admin_count = db.execute("SELECT COUNT(*) c FROM users WHERE role='admin'").fetchone()["c"]
        if admin_count <= 1:
            return jsonify(ok=False, msg="Cannot change role: this is the only admin account."), 400

    db.execute("UPDATE users SET name=?, phone=?, designation=?, role=? WHERE emp_id=?",
               (name, phone, desg, role, emp_id))
    db.commit()
    return jsonify(ok=True, msg="Employee updated.")


@app.route("/api/admin/reset-password", methods=["POST"])
@admin_required
def api_admin_reset_password():
    """Admin sets a temporary password for an employee."""
    d = request.get_json(force=True)
    emp_id = (d.get("emp_id") or "").strip()
    new_pw = d.get("new_password") or ""
    if not emp_id or not new_pw:
        return jsonify(ok=False, msg="Missing employee ID or new password."), 400
    if len(new_pw) < 6:
        return jsonify(ok=False, msg="Temporary password must be at least 6 characters."), 400

    db = get_db()
    target = db.execute("SELECT 1 FROM users WHERE emp_id = ?", (emp_id,)).fetchone()
    if target is None:
        return jsonify(ok=False, msg="Employee not found."), 404

    db.execute("UPDATE users SET password_hash=?, must_reset=1 WHERE emp_id=?",
               (generate_password_hash(new_pw), emp_id))
    db.commit()
    return jsonify(ok=True, msg="Password reset. Share the temporary password with the employee.")


@app.route("/api/admin/delete-user", methods=["POST"])
@admin_required
def api_admin_delete_user():
    """Delete an employee (and their scores). Admin accounts are protected."""
    d = request.get_json(force=True)
    emp_id = (d.get("emp_id") or "").strip()
    if not emp_id:
        return jsonify(ok=False, msg="Missing employee ID."), 400

    db = get_db()
    target = db.execute("SELECT * FROM users WHERE emp_id = ?", (emp_id,)).fetchone()
    if target is None:
        return jsonify(ok=False, msg="Employee not found."), 404
    if target["role"] == "admin":
        return jsonify(ok=False, msg="Admin accounts cannot be deleted here."), 400

    db.execute("DELETE FROM users WHERE emp_id = ?", (emp_id,))
    db.execute("DELETE FROM scores WHERE emp_id = ?", (emp_id,))
    db.commit()
    return jsonify(ok=True, msg="Employee deleted.")


# ---------- SETTINGS: team overview, admins/instructors, quick add, backup ----------
@app.route("/api/admin/team-overview")
@admin_required
def api_admin_team_overview():
    db = get_db()
    def cnt(q, *a): return db.execute(q, a).fetchone()["c"]
    total = cnt("SELECT COUNT(*) c FROM users")
    learners = cnt("SELECT COUNT(*) c FROM users WHERE role='staff'")
    instructors = cnt("SELECT COUNT(*) c FROM users WHERE role='instructor'")
    admins = cnt("SELECT COUNT(*) c FROM users WHERE role='admin'")
    pending = cnt("SELECT COUNT(*) c FROM users WHERE status='pending'")
    admin_rows = db.execute("SELECT emp_id,name,designation,status FROM users WHERE role='admin' ORDER BY name").fetchall()
    inst_rows = db.execute("SELECT emp_id,name,designation,status FROM users WHERE role='instructor' ORDER BY name").fetchall()
    return jsonify(ok=True,
                   counts={"total": total, "learners": learners, "instructors": instructors,
                           "admins": admins, "pending": pending},
                   admins=[dict(r) for r in admin_rows],
                   instructors=[dict(r) for r in inst_rows])


@app.route("/api/admin/quick-add-user", methods=["POST"])
@admin_required
def api_admin_quick_add_user():
    """Add a single user directly (name, emp_id, phone, designation, role, password)."""
    d = request.get_json(force=True)
    name = (d.get("name") or "").strip()
    emp_id = (d.get("emp_id") or "").strip()
    phone = (d.get("phone") or "").strip()
    desg = (d.get("designation") or "").strip()
    role = (d.get("role") or "staff").strip()
    pw = d.get("password") or ""
    auto_approve = d.get("auto_approve", True)

    if not name or not emp_id:
        return jsonify(ok=False, msg="Name and Employee ID are required."), 400
    if role not in VALID_ROLES:
        role = "staff"
    if not pw:
        pw = "Golisoda@123"
    if len(pw) < 6:
        return jsonify(ok=False, msg="Password must be at least 6 characters."), 400

    db = get_db()
    if db.execute("SELECT 1 FROM users WHERE emp_id=?", (emp_id,)).fetchone():
        return jsonify(ok=False, msg="This Employee ID already exists."), 400

    status = "approved" if auto_approve else "pending"
    must_reset = 1 if pw == "Golisoda@123" else 0
    db.execute(
        "INSERT INTO users (emp_id,name,phone,designation,password_hash,status,role,created_at,must_reset) "
        "VALUES (?,?,?,?,?,?,?,?,?)",
        (emp_id, name, phone, desg, generate_password_hash(pw), status, role,
         datetime.utcnow().isoformat(), must_reset)
    )
    db.commit()
    return jsonify(ok=True, msg=f"{name} added as {role}.")


@app.route("/api/admin/set-admin", methods=["POST"])
@admin_required
def api_admin_set_admin():
    """Promote a user to admin, or remove admin rights (back to staff)."""
    d = request.get_json(force=True)
    emp_id = (d.get("emp_id") or "").strip()
    make_admin = d.get("make_admin", True)
    db = get_db()
    target = db.execute("SELECT * FROM users WHERE emp_id=?", (emp_id,)).fetchone()
    if not target:
        return jsonify(ok=False, msg="User not found."), 404
    if not make_admin and target["role"] == "admin":
        admin_count = db.execute("SELECT COUNT(*) c FROM users WHERE role='admin'").fetchone()["c"]
        if admin_count <= 1:
            return jsonify(ok=False, msg="Cannot remove the only admin."), 400
    new_role = "admin" if make_admin else "staff"
    db.execute("UPDATE users SET role=?, status='approved' WHERE emp_id=?", (new_role, emp_id))
    db.commit()
    return jsonify(ok=True, msg=("Now an admin." if make_admin else "Admin rights removed."))


@app.route("/api/admin/export-data")
@admin_required
def api_admin_export_data():
    """Export all users + scores + assessment results as a single CSV (backup)."""
    db = get_db()
    out = io.StringIO()
    w = csv.writer(out)

    w.writerow(["=== EMPLOYEES ==="])
    w.writerow(["emp_id", "name", "phone", "designation", "role", "status", "created_at", "last_login"])
    for u in db.execute("SELECT emp_id,name,phone,designation,role,status,created_at,last_login FROM users ORDER BY name").fetchall():
        w.writerow([u["emp_id"], u["name"], u["phone"], u["designation"], u["role"], u["status"], u["created_at"], u["last_login"]])

    w.writerow([])
    w.writerow(["=== MODULE QUIZ SCORES ==="])
    w.writerow(["emp_id", "module_id", "percent", "passed", "taken_at"])
    for s in db.execute("SELECT emp_id,module_id,percent,passed,taken_at FROM scores ORDER BY taken_at").fetchall():
        w.writerow([s["emp_id"], s["module_id"], s["percent"], s["passed"], s["taken_at"]])

    w.writerow([])
    w.writerow(["=== ASSESSMENT RESULTS ==="])
    w.writerow(["emp_id", "assessment_id", "percent", "passed", "taken_at"])
    for r in db.execute("SELECT emp_id,assessment_id,percent,passed,taken_at FROM assessment_results ORDER BY taken_at").fetchall():
        w.writerow([r["emp_id"], r["assessment_id"], r["percent"], r["passed"], r["taken_at"]])

    from flask import Response
    csv_data = out.getvalue()
    fname = "golisoda_backup_" + datetime.utcnow().strftime("%Y%m%d_%H%M") + ".csv"
    return Response(csv_data, mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename={fname}"})


@app.route("/api/admin/bulk-add", methods=["POST"])
@admin_required
def api_admin_bulk_add():
    """Bulk add employees from CSV text.
    Expected columns: name, emp_id, phone, designation, password (optional).
    If password missing, a default is set and must_reset flagged."""
    d = request.get_json(force=True)
    csv_text = d.get("csv") or ""
    default_status = "approved" if d.get("auto_approve") else "pending"
    if not csv_text.strip():
        return jsonify(ok=False, msg="No CSV content received."), 400

    db = get_db()
    reader = csv.DictReader(io.StringIO(csv_text))
    added, skipped, errors = 0, 0, []

    for i, raw in enumerate(reader, start=2):
        row = { (k or "").strip().lower(): (v or "").strip() for k, v in raw.items() }
        name = row.get("name", "")
        emp_id = row.get("emp_id") or row.get("employee_id") or row.get("emp id", "")
        phone = row.get("phone", "")
        desg = row.get("designation") or row.get("role_title", "")
        pw = row.get("password", "") or "Golisoda@123"

        if not name or not emp_id:
            errors.append(f"Row {i}: missing name or emp_id")
            continue
        exists = db.execute("SELECT 1 FROM users WHERE emp_id = ?", (emp_id,)).fetchone()
        if exists:
            skipped += 1
            continue
        db.execute(
            "INSERT INTO users (emp_id,name,phone,designation,password_hash,status,role,created_at,must_reset) "
            "VALUES (?,?,?,?,?,?,?,?,?)",
            (emp_id, name, phone, desg, generate_password_hash(pw),
             default_status, "staff", datetime.utcnow().isoformat(), 1)
        )
        added += 1

    db.commit()
    return jsonify(ok=True, added=added, skipped=skipped, errors=errors,
                   msg=f"Added {added}, skipped {skipped} (already existed).")


# ===============================================================
#  NEW — DELIVERY 3: Assessment Engine
# ===============================================================
import random as _random

FALLBACK_ROLES = ["BDE", "BDM", "State Head", "RSM", "NSM", "Corporate", "Back Office"]


def get_role_choices():
    """Designations, read live from the database so admins can add their own."""
    try:
        rows = get_db().execute(
            "SELECT name FROM designations ORDER BY sort_order, name"
        ).fetchall()
        return [r["name"] for r in rows] or FALLBACK_ROLES
    except Exception:
        return FALLBACK_ROLES


def _user_role_label(u):
    """Map a user to the assessment role label using their designation."""
    return (u["designation"] or "").strip()


def _assessment_allowed_for(assessment, u):
    """Check if this user may take this assessment based on roles field."""
    roles = (assessment["roles"] or "all").strip().lower()
    if roles in ("", "all"):
        return True
    desg = (_user_role_label(u) or "").lower()
    allowed = [r.strip().lower() for r in roles.split(",")]
    return any(a and a in desg for a in allowed)


def _roles_match_designation(roles_field, designation):
    """True if a user's designation matches a roles field (comma list or 'all')."""
    roles = (roles_field or "all").strip().lower()
    if roles in ("", "all"):
        return True
    desg = (designation or "").strip().lower()
    allowed = [r.strip().lower() for r in roles.split(",")]
    return any(a and a in desg for a in allowed)


@app.route("/api/admin/completion-by-person")
@admin_required
def api_admin_completion_by_person():
    """Per-employee pending counts, grouped by designation (for the folders view).
    For each approved learner: how many induction/training modules and
    assessments they still have pending, matched to their designation."""
    db = get_db()
    learners = db.execute(
        "SELECT emp_id, name, designation FROM users "
        "WHERE status='approved' AND role='staff' ORDER BY designation, name"
    ).fetchall()

    ind = db.execute("SELECT id, roles FROM content_modules WHERE kind='induction' AND status='live'").fetchall()
    trn = db.execute("SELECT id, roles FROM content_modules WHERE kind='training' AND status='live'").fetchall()
    ass = db.execute("SELECT id, roles FROM assessments WHERE active=1").fetchall()
    vids = db.execute("SELECT id, roles FROM videos WHERE status='live'").fetchall()

    comp = {}
    for c in db.execute("SELECT module_id, emp_id FROM module_completions").fetchall():
        comp.setdefault(c["emp_id"], set()).add(c["module_id"])
    passed = {}
    for pr in db.execute("SELECT assessment_id, emp_id FROM assessment_results WHERE passed=1").fetchall():
        passed.setdefault(pr["emp_id"], set()).add(pr["assessment_id"])
    vwatch = {}
    for w in db.execute("SELECT video_id, emp_id FROM video_completions").fetchall():
        vwatch.setdefault(w["emp_id"], set()).add(w["video_id"])

    # start every known designation as an empty folder, so all of them show
    groups = {}
    try:
        for dr in db.execute("SELECT name FROM designations ORDER BY sort_order, name").fetchall():
            groups[dr["name"]] = []
    except Exception:
        pass

    for u in learners:
        desg = (u["designation"] or "").strip() or "No designation"
        mine_i = [m for m in ind if _roles_match_designation(m["roles"], u["designation"])]
        mine_t = [m for m in trn if _roles_match_designation(m["roles"], u["designation"])]
        mine_a = [a for a in ass if _roles_match_designation(a["roles"], u["designation"])]
        mine_v = [v for v in vids if _roles_match_designation(v["roles"], u["designation"])]
        done = comp.get(u["emp_id"], set())
        pa = passed.get(u["emp_id"], set())
        vw = vwatch.get(u["emp_id"], set())
        pi = sum(1 for m in mine_i if m["id"] not in done)
        pt = sum(1 for m in mine_t if m["id"] not in done)
        pas = sum(1 for a in mine_a if a["id"] not in pa)
        pv = sum(1 for v in mine_v if v["id"] not in vw)
        total_pending = pi + pt + pas + pv
        total_items = len(mine_i) + len(mine_t) + len(mine_a) + len(mine_v)
        groups.setdefault(desg, []).append({
            "name": u["name"], "emp_id": u["emp_id"],
            "induction_pending": pi, "training_pending": pt,
            "assess_pending": pas, "video_pending": pv,
            "total_pending": total_pending, "total_items": total_items
        })

    out = []
    for desg, people in groups.items():
        pending_people = sum(1 for p in people if p["total_pending"] > 0)
        out.append({"designation": desg, "count": len(people),
                    "pending_people": pending_people, "people": people})
    return jsonify(ok=True, groups=out, total_learners=len(learners))


@app.route("/api/admin/completion-report")
@admin_required
def api_admin_completion_report():
    """For every assessment and training/induction module, show enrolled /
    completed / pending counts + the list of people who haven't done it.
    Enrolled = approved learners whose designation matches the item's roles.
    Assessment completed = passed (score >= pass mark).
    Module completed = has a module_completions record."""
    db = get_db()

    # all approved learners (staff role)
    learners = db.execute(
        "SELECT emp_id, name, designation FROM users "
        "WHERE status='approved' AND role='staff'"
    ).fetchall()

    # ---- ASSESSMENTS ----
    assessments = db.execute("SELECT * FROM assessments ORDER BY roles, title").fetchall()
    # who passed which assessment
    passed_rows = db.execute(
        "SELECT DISTINCT assessment_id, emp_id FROM assessment_results WHERE passed=1"
    ).fetchall()
    passed_set = {(r["assessment_id"], r["emp_id"]) for r in passed_rows}

    assess_report = []
    for a in assessments:
        enrolled = [u for u in learners if _roles_match_designation(a["roles"], u["designation"])]
        done, pending = [], []
        for u in enrolled:
            if (a["id"], u["emp_id"]) in passed_set:
                done.append(u)
            else:
                pending.append(u)
        assess_report.append({
            "id": a["id"], "title": a["title"], "roles": a["roles"],
            "enrolled": len(enrolled), "completed": len(done), "pending": len(pending),
            "pending_people": [{"name": p["name"], "emp_id": p["emp_id"],
                                "designation": p["designation"]} for p in pending]
        })

    # ---- MODULES (induction + training) ----
    modules = db.execute(
        "SELECT * FROM content_modules WHERE status='live' ORDER BY kind, sort_order, id"
    ).fetchall()
    comp_rows = db.execute("SELECT DISTINCT module_id, emp_id FROM module_completions").fetchall()
    comp_set = {(r["module_id"], r["emp_id"]) for r in comp_rows}

    module_report = []
    for m in modules:
        enrolled = [u for u in learners if _roles_match_designation(m["roles"], u["designation"])]
        done, pending = [], []
        for u in enrolled:
            if (m["id"], u["emp_id"]) in comp_set:
                done.append(u)
            else:
                pending.append(u)
        module_report.append({
            "id": m["id"], "title": m["title"], "kind": m["kind"], "roles": m["roles"],
            "enrolled": len(enrolled), "completed": len(done), "pending": len(pending),
            "pending_people": [{"name": p["name"], "emp_id": p["emp_id"],
                                "designation": p["designation"]} for p in pending]
        })

    return jsonify(ok=True, assessments=assess_report, modules=module_report,
                   total_learners=len(learners))


@app.route("/api/admin/completion-report.csv")
@admin_required
def api_admin_completion_report_csv():
    """Download the completion report as CSV."""
    # rebuild the same data
    import json as _json
    from flask import Response
    with app.test_request_context():
        pass
    db = get_db()
    learners = db.execute(
        "SELECT emp_id, name, designation FROM users WHERE status='approved' AND role='staff'"
    ).fetchall()
    passed_set = {(r["assessment_id"], r["emp_id"]) for r in db.execute(
        "SELECT DISTINCT assessment_id, emp_id FROM assessment_results WHERE passed=1").fetchall()}
    comp_set = {(r["module_id"], r["emp_id"]) for r in db.execute(
        "SELECT DISTINCT module_id, emp_id FROM module_completions").fetchall()}

    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Type", "Title", "Roles", "Enrolled", "Completed", "Pending", "Pending People (name - id)"])

    for a in db.execute("SELECT * FROM assessments ORDER BY roles, title").fetchall():
        enrolled = [u for u in learners if _roles_match_designation(a["roles"], u["designation"])]
        pending = [u for u in enrolled if (a["id"], u["emp_id"]) not in passed_set]
        names = "; ".join(f"{p['name']} - {p['emp_id']}" for p in pending)
        w.writerow(["Assessment", a["title"], a["roles"], len(enrolled),
                    len(enrolled) - len(pending), len(pending), names])

    for m in db.execute("SELECT * FROM content_modules WHERE status='live' ORDER BY kind, sort_order, id").fetchall():
        enrolled = [u for u in learners if _roles_match_designation(m["roles"], u["designation"])]
        pending = [u for u in enrolled if (m["id"], u["emp_id"]) not in comp_set]
        names = "; ".join(f"{p['name']} - {p['emp_id']}" for p in pending)
        w.writerow([f"Module ({m['kind']})", m["title"], m["roles"], len(enrolled),
                    len(enrolled) - len(pending), len(pending), names])

    return Response(out.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": 'attachment; filename="completion_report.csv"'})


@app.route("/api/admin/completion-report.xlsx")
@admin_required
def api_admin_completion_report_xlsx():
    """Download the completion report as Excel."""
    from flask import Response
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except Exception:
        return jsonify(ok=False, msg="Excel export not available."), 500

    db = get_db()
    learners = db.execute(
        "SELECT emp_id, name, designation FROM users WHERE status='approved' AND role='staff'"
    ).fetchall()
    passed_set = {(r["assessment_id"], r["emp_id"]) for r in db.execute(
        "SELECT DISTINCT assessment_id, emp_id FROM assessment_results WHERE passed=1").fetchall()}
    comp_set = {(r["module_id"], r["emp_id"]) for r in db.execute(
        "SELECT DISTINCT module_id, emp_id FROM module_completions").fetchall()}

    wb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="00AEEF")

    def _fill_sheet(ws, rows_source, is_assessment):
        ws.append(["Title", "Roles", "Enrolled", "Completed", "Pending", "Pending People"])
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = hdr_fill
        for it in rows_source:
            enrolled = [u for u in learners if _roles_match_designation(it["roles"], u["designation"])]
            key_set = passed_set if is_assessment else comp_set
            pending = [u for u in enrolled if (it["id"], u["emp_id"]) not in key_set]
            names = "; ".join(f"{p['name']} ({p['emp_id']})" for p in pending)
            ws.append([it["title"], it["roles"], len(enrolled),
                       len(enrolled) - len(pending), len(pending), names])
        for i, wdt in enumerate([34, 20, 10, 11, 9, 60], start=1):
            ws.column_dimensions[chr(64 + i)].width = wdt

    ws1 = wb.active; ws1.title = "Assessments"
    _fill_sheet(ws1, db.execute("SELECT * FROM assessments ORDER BY roles, title").fetchall(), True)

    ws2 = wb.create_sheet("Modules")
    _fill_sheet(ws2, db.execute("SELECT * FROM content_modules WHERE status='live' ORDER BY kind, sort_order, id").fetchall(), False)

    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return Response(bio.getvalue(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": 'attachment; filename="completion_report.xlsx"'})


def _designation_usage(name):
    """How many employees and items currently use this designation."""
    db = get_db()
    low = (name or "").strip().lower()
    emp = db.execute(
        "SELECT COUNT(*) c FROM users WHERE LOWER(TRIM(designation)) = ?", (low,)
    ).fetchone()["c"]

    def _count(table):
        n = 0
        for r in db.execute(f"SELECT roles FROM {table}").fetchall():
            parts = [p.strip().lower() for p in (r["roles"] or "").split(",")]
            if low in parts:
                n += 1
        return n

    return {"employees": emp,
            "assessments": _count("assessments"),
            "modules": _count("content_modules")}


@app.route("/api/admin/designations")
@admin_required
def api_admin_designations():
    """List designations with how many employees/items use each."""
    db = get_db()
    rows = db.execute("SELECT * FROM designations ORDER BY sort_order, name").fetchall()
    out = []
    for r in rows:
        u = _designation_usage(r["name"])
        out.append({"id": r["id"], "name": r["name"], **u})
    return jsonify(ok=True, designations=out)


@app.route("/api/admin/save-designation", methods=["POST"])
@admin_required
def api_admin_save_designation():
    """Add a new designation, or rename an existing one.
    A rename cascades: every employee on the old name is updated, and the
    old name is swapped inside assessment/module role lists."""
    d = request.get_json(force=True)
    did = d.get("id")
    name = (d.get("name") or "").strip()
    if not name:
        return jsonify(ok=False, msg="Please enter a designation name.")

    db = get_db()
    clash = db.execute(
        "SELECT id FROM designations WHERE LOWER(name) = ? AND id != ?",
        (name.lower(), did or -1)
    ).fetchone()
    if clash:
        return jsonify(ok=False, msg=f'"{name}" already exists.')

    if did:
        old = db.execute("SELECT name FROM designations WHERE id=?", (did,)).fetchone()
        if not old:
            return jsonify(ok=False, msg="Designation not found.")
        old_name = old["name"]
        db.execute("UPDATE designations SET name=? WHERE id=?", (name, did))

        moved = 0
        if old_name.strip().lower() != name.strip().lower():
            # 1) every employee carrying the old designation
            cur = db.execute(
                "UPDATE users SET designation=? WHERE LOWER(TRIM(designation))=?",
                (name, old_name.strip().lower())
            )
            moved = cur.rowcount if hasattr(cur, "rowcount") else 0

            # 2) the old name inside comma-separated role lists
            for table in ("assessments", "content_modules", "certificate_tracks"):
                try:
                    for r in db.execute(f"SELECT id, roles FROM {table}").fetchall():
                        parts = [p.strip() for p in (r["roles"] or "").split(",")]
                        if any(p.lower() == old_name.strip().lower() for p in parts):
                            new_parts = [name if p.lower() == old_name.strip().lower() else p
                                         for p in parts if p]
                            db.execute(f"UPDATE {table} SET roles=? WHERE id=?",
                                       (",".join(new_parts), r["id"]))
                except Exception:
                    pass  # table may not have a roles column

        db.commit()
        msg = f'Renamed to "{name}".'
        if moved:
            msg += f" {moved} employee(s) updated."
        return jsonify(ok=True, msg=msg)

    nxt = db.execute("SELECT COALESCE(MAX(sort_order), 0) + 1 AS n FROM designations").fetchone()["n"]
    db.execute("INSERT INTO designations (name, sort_order, created_at) VALUES (?,?,?)",
               (name, nxt, datetime.utcnow().isoformat()))
    db.commit()
    return jsonify(ok=True, msg=f'"{name}" added.')


@app.route("/api/admin/delete-designation", methods=["POST"])
@admin_required
def api_admin_delete_designation():
    """Delete a designation. Reports usage first unless force=true."""
    d = request.get_json(force=True)
    did = d.get("id")
    force = bool(d.get("force"))

    db = get_db()
    row = db.execute("SELECT name FROM designations WHERE id=?", (did,)).fetchone()
    if not row:
        return jsonify(ok=False, msg="Designation not found.")

    use = _designation_usage(row["name"])
    total = use["employees"] + use["assessments"] + use["modules"]
    if total and not force:
        bits = []
        if use["employees"]:
            bits.append(f"{use['employees']} employee(s)")
        if use["assessments"]:
            bits.append(f"{use['assessments']} assessment(s)")
        if use["modules"]:
            bits.append(f"{use['modules']} module(s)")
        return jsonify(ok=False, in_use=True, usage=use,
                       msg=f'"{row["name"]}" is being used by ' + ", ".join(bits) +
                           ". Rename it instead, or confirm to delete it anyway.")

    db.execute("DELETE FROM designations WHERE id=?", (did,))
    db.commit()
    return jsonify(ok=True, msg=f'"{row["name"]}" deleted.')


@app.route("/api/admin/assessments")
@admin_required
def api_admin_assessments():
    """List all assessments with question counts and attempt counts."""
    db = get_db()
    rows = db.execute("SELECT * FROM assessments ORDER BY created_at DESC").fetchall()
    out = []
    for a in rows:
        qn = db.execute("SELECT COUNT(*) c FROM questions WHERE assessment_id=?", (a["id"],)).fetchone()["c"]
        at = db.execute("SELECT COUNT(*) c FROM assessment_results WHERE assessment_id=?", (a["id"],)).fetchone()["c"]
        d = dict(a); d["question_count"] = qn; d["attempt_count"] = at
        out.append(d)
    return jsonify(ok=True, assessments=out, role_choices=get_role_choices())


@app.route("/api/admin/create-assessment", methods=["POST"])
@admin_required
def api_admin_create_assessment():
    """Create an assessment and load its question pool from CSV."""
    d = request.get_json(force=True)
    title = (d.get("title") or "").strip()
    desc = (d.get("description") or "").strip()
    roles = (d.get("roles") or "all").strip() or "all"
    csv_text = d.get("csv") or ""
    try:
        num_q = int(d.get("num_questions") or 10)
    except (ValueError, TypeError):
        num_q = 10
    try:
        pass_pct = int(d.get("pass_percent") or 90)
    except (ValueError, TypeError):
        pass_pct = 90
    if pass_pct < 90:
        pass_pct = 90  # enforce the "serious" minimum
    if pass_pct > 100:
        pass_pct = 100
    try:
        time_limit = int(d.get("time_limit") or 0)
    except (ValueError, TypeError):
        time_limit = 0

    if not title:
        return jsonify(ok=False, msg="Please give the assessment a title."), 400
    if not csv_text.strip():
        return jsonify(ok=False, msg="Please paste the question CSV."), 400

    # Parse questions
    reader = csv.DictReader(io.StringIO(csv_text))
    parsed, errors = [], []
    for i, raw in enumerate(reader, start=2):
        row = {(k or "").strip().lower(): (v or "").strip() for k, v in raw.items()}
        q = row.get("question", "")
        a = row.get("option_a") or row.get("opt_a") or row.get("a", "")
        b = row.get("option_b") or row.get("opt_b") or row.get("b", "")
        cc = row.get("option_c") or row.get("opt_c") or row.get("c", "")
        dd = row.get("option_d") or row.get("opt_d") or row.get("d", "")
        correct = (row.get("correct") or row.get("answer") or "").strip().upper()
        cat = row.get("category", "")
        if not q or not a or not b:
            errors.append(f"Row {i}: missing question or options A/B")
            continue
        if correct not in ("A", "B", "C", "D"):
            errors.append(f"Row {i}: 'correct' must be A, B, C, or D (got '{correct}')")
            continue
        if correct == "C" and not cc:
            errors.append(f"Row {i}: correct is C but option C is empty")
            continue
        if correct == "D" and not dd:
            errors.append(f"Row {i}: correct is D but option D is empty")
            continue
        parsed.append((q, a, b, cc, dd, correct, cat))

    if not parsed:
        return jsonify(ok=False, msg="No valid questions found.", errors=errors), 400

    if num_q > len(parsed):
        num_q = len(parsed)

    db = get_db()
    u = current_user()
    cur = db.execute(
        "INSERT INTO assessments (title,description,roles,num_questions,pass_percent,time_limit,active,created_by,created_at) "
        "VALUES (?,?,?,?,?,?,?,?,?) RETURNING id",
        (title, desc, roles, num_q, pass_pct, time_limit, 1, u["emp_id"], datetime.utcnow().isoformat())
    )
    aid = cur.fetchone()["id"]
    for (q, a, b, cc, dd, correct, cat) in parsed:
        db.execute(
            "INSERT INTO questions (assessment_id,question,opt_a,opt_b,opt_c,opt_d,correct,category) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (aid, q, a, b, cc, dd, correct, cat)
        )
    db.commit()
    return jsonify(ok=True, id=aid, loaded=len(parsed), errors=errors,
                   msg=f"Assessment created with {len(parsed)} questions.")


@app.route("/api/admin/toggle-assessment", methods=["POST"])
@admin_required
def api_admin_toggle_assessment():
    d = request.get_json(force=True)
    aid = d.get("id")
    db = get_db()
    a = db.execute("SELECT active FROM assessments WHERE id=?", (aid,)).fetchone()
    if not a:
        return jsonify(ok=False, msg="Not found."), 404
    newv = 0 if a["active"] else 1
    db.execute("UPDATE assessments SET active=? WHERE id=?", (newv, aid))
    db.commit()
    return jsonify(ok=True, active=bool(newv))


@app.route("/api/admin/delete-assessment", methods=["POST"])
@admin_required
def api_admin_delete_assessment():
    d = request.get_json(force=True)
    aid = d.get("id")
    db = get_db()
    db.execute("DELETE FROM questions WHERE assessment_id=?", (aid,))
    db.execute("DELETE FROM assessment_results WHERE assessment_id=?", (aid,))
    db.execute("DELETE FROM assessments WHERE id=?", (aid,))
    db.commit()
    return jsonify(ok=True, msg="Assessment deleted.")


# ---------- EDIT an assessment: settings + questions ----------
@app.route("/api/admin/assessment-detail")
@admin_required
def api_admin_assessment_detail():
    """Full assessment + all its questions (for the edit screen)."""
    aid = request.args.get("id")
    db = get_db()
    a = db.execute("SELECT * FROM assessments WHERE id=?", (aid,)).fetchone()
    if not a:
        return jsonify(ok=False, msg="Not found."), 404
    qs = db.execute("SELECT * FROM questions WHERE assessment_id=? ORDER BY id", (aid,)).fetchall()
    return jsonify(ok=True, assessment=dict(a), questions=[dict(q) for q in qs],
                   role_choices=get_role_choices())


@app.route("/api/admin/update-assessment", methods=["POST"])
@admin_required
def api_admin_update_assessment():
    """Update assessment settings (title, roles, pass mark, num questions, time)."""
    d = request.get_json(force=True)
    aid = d.get("id")
    db = get_db()
    a = db.execute("SELECT * FROM assessments WHERE id=?", (aid,)).fetchone()
    if not a:
        return jsonify(ok=False, msg="Not found."), 404
    title = (d.get("title") or a["title"]).strip()
    roles = (d.get("roles") or "all").strip() or "all"
    try: num_q = int(d.get("num_questions") or a["num_questions"])
    except (ValueError, TypeError): num_q = a["num_questions"]
    try: pass_pct = int(d.get("pass_percent") or a["pass_percent"])
    except (ValueError, TypeError): pass_pct = a["pass_percent"]
    if pass_pct < 90: pass_pct = 90
    if pass_pct > 100: pass_pct = 100
    try: time_limit = int(d.get("time_limit") or 0)
    except (ValueError, TypeError): time_limit = 0

    # don't allow num_q bigger than the pool
    pool = db.execute("SELECT COUNT(*) c FROM questions WHERE assessment_id=?", (aid,)).fetchone()["c"]
    if num_q > pool and pool > 0:
        num_q = pool

    db.execute("UPDATE assessments SET title=?,roles=?,num_questions=?,pass_percent=?,time_limit=? WHERE id=?",
               (title, roles, num_q, pass_pct, time_limit, aid))
    db.commit()
    return jsonify(ok=True, msg="Settings saved.")


@app.route("/api/admin/save-question", methods=["POST"])
@admin_required
def api_admin_save_question():
    """Add a new question or update an existing one."""
    d = request.get_json(force=True)
    aid = d.get("assessment_id")
    qid = d.get("id")  # None = new question
    q = (d.get("question") or "").strip()
    a = (d.get("opt_a") or "").strip()
    b = (d.get("opt_b") or "").strip()
    cc = (d.get("opt_c") or "").strip()
    dd = (d.get("opt_d") or "").strip()
    correct = (d.get("correct") or "").strip().upper()
    cat = (d.get("category") or "").strip()

    if not q or not a or not b:
        return jsonify(ok=False, msg="Question and options A and B are required."), 400
    if correct not in ("A", "B", "C", "D"):
        return jsonify(ok=False, msg="Correct answer must be A, B, C, or D."), 400
    if correct == "C" and not cc:
        return jsonify(ok=False, msg="Correct is C but option C is empty."), 400
    if correct == "D" and not dd:
        return jsonify(ok=False, msg="Correct is D but option D is empty."), 400

    db = get_db()
    if qid:
        db.execute("UPDATE questions SET question=?,opt_a=?,opt_b=?,opt_c=?,opt_d=?,correct=?,category=? WHERE id=? AND assessment_id=?",
                   (q, a, b, cc, dd, correct, cat, qid, aid))
        msg = "Question updated."
    else:
        db.execute("INSERT INTO questions (assessment_id,question,opt_a,opt_b,opt_c,opt_d,correct,category) VALUES (?,?,?,?,?,?,?,?)",
                   (aid, q, a, b, cc, dd, correct, cat))
        msg = "Question added."
    db.commit()
    return jsonify(ok=True, msg=msg)


@app.route("/api/admin/delete-question", methods=["POST"])
@admin_required
def api_admin_delete_question():
    """Remove a single question from an assessment."""
    d = request.get_json(force=True)
    qid = d.get("id")
    aid = d.get("assessment_id")
    db = get_db()
    db.execute("DELETE FROM questions WHERE id=? AND assessment_id=?", (qid, aid))
    # if num_questions now exceeds pool, shrink it
    pool = db.execute("SELECT COUNT(*) c FROM questions WHERE assessment_id=?", (aid,)).fetchone()["c"]
    a = db.execute("SELECT num_questions FROM assessments WHERE id=?", (aid,)).fetchone()
    if a and pool > 0 and a["num_questions"] > pool:
        db.execute("UPDATE assessments SET num_questions=? WHERE id=?", (pool, aid))
    db.commit()
    return jsonify(ok=True, msg="Question removed.", pool=pool)


@app.route("/api/admin/add-questions-csv", methods=["POST"])
@admin_required
def api_admin_add_questions_csv():
    """Append more questions to an existing assessment from CSV."""
    d = request.get_json(force=True)
    aid = d.get("id")
    csv_text = d.get("csv") or ""
    db = get_db()
    a = db.execute("SELECT * FROM assessments WHERE id=?", (aid,)).fetchone()
    if not a:
        return jsonify(ok=False, msg="Not found."), 404
    reader = csv.DictReader(io.StringIO(csv_text))
    parsed, errors = [], []
    for i, raw in enumerate(reader, start=2):
        row = {(k or "").strip().lower(): (v or "").strip() for k, v in raw.items()}
        q = row.get("question", "")
        aa = row.get("option_a") or row.get("opt_a") or row.get("a", "")
        bb = row.get("option_b") or row.get("opt_b") or row.get("b", "")
        cc = row.get("option_c") or row.get("opt_c") or row.get("c", "")
        dd = row.get("option_d") or row.get("opt_d") or row.get("d", "")
        correct = (row.get("correct") or row.get("answer") or "").strip().upper()
        cat = row.get("category", "")
        if not q or not aa or not bb:
            errors.append(f"Row {i}: missing question or options A/B"); continue
        if correct not in ("A", "B", "C", "D"):
            errors.append(f"Row {i}: correct must be A/B/C/D"); continue
        parsed.append((q, aa, bb, cc, dd, correct, cat))
    if not parsed:
        return jsonify(ok=False, msg="No valid questions found.", errors=errors), 400
    for (q, aa, bb, cc, dd, correct, cat) in parsed:
        db.execute("INSERT INTO questions (assessment_id,question,opt_a,opt_b,opt_c,opt_d,correct,category) VALUES (?,?,?,?,?,?,?,?)",
                   (aid, q, aa, bb, cc, dd, correct, cat))
    db.commit()
    return jsonify(ok=True, msg=f"Added {len(parsed)} more questions.", added=len(parsed), errors=errors)


@app.route("/api/admin/assessment-results")
@admin_required
def api_admin_assessment_results():
    """All results for one assessment (for the admin table + CSV export)."""
    aid = request.args.get("id")
    db = get_db()
    rows = db.execute(
        "SELECT r.*, u.name, u.designation FROM assessment_results r "
        "LEFT JOIN users u ON u.emp_id = r.emp_id "
        "WHERE r.assessment_id=? ORDER BY r.taken_at DESC", (aid,)
    ).fetchall()
    return jsonify(ok=True, results=[dict(r) for r in rows])


def _fmt_dt(iso):
    """Human-friendly date+time from an ISO string."""
    try:
        dt = datetime.fromisoformat((iso or "").replace("Z", ""))
        return dt.strftime("%d %b %Y, %I:%M %p")
    except Exception:
        return iso or ""


def _fmt_mins(seconds):
    try:
        s = int(seconds or 0)
    except (ValueError, TypeError):
        s = 0
    if s <= 0:
        return ""
    m, sec = divmod(s, 60)
    return f"{m}m {sec}s" if m else f"{sec}s"


@app.route("/api/admin/attempt-details")
@admin_required
def api_admin_attempt_details():
    """Full question-by-question breakdown for ONE attempt (on-screen view)."""
    result_id = request.args.get("result_id")
    db = get_db()
    r = db.execute(
        "SELECT r.*, u.name, u.designation, a.title FROM assessment_results r "
        "LEFT JOIN users u ON u.emp_id = r.emp_id "
        "LEFT JOIN assessments a ON a.id = r.assessment_id "
        "WHERE r.id=?", (result_id,)
    ).fetchone()
    if not r:
        return jsonify(ok=False, msg="Attempt not found."), 404
    details = db.execute(
        "SELECT * FROM answer_details WHERE result_id=? ORDER BY id", (result_id,)
    ).fetchall()
    return jsonify(ok=True,
                   header={
                       "name": r["name"], "emp_id": r["emp_id"], "designation": r["designation"],
                       "assessment": r["title"], "score": r["score"], "total": r["total"],
                       "percent": r["percent"], "passed": bool(r["passed"]),
                       "taken_at": _fmt_dt(r["taken_at"]), "time_taken": _fmt_mins(r["time_taken"])
                   },
                   details=[dict(d) for d in details])


def _gather_report_rows(aid):
    """Return (assessment, attempts, all_details) for building a report."""
    db = get_db()
    a = db.execute("SELECT * FROM assessments WHERE id=?", (aid,)).fetchone()
    attempts = db.execute(
        "SELECT r.*, u.name, u.designation FROM assessment_results r "
        "LEFT JOIN users u ON u.emp_id = r.emp_id "
        "WHERE r.assessment_id=? ORDER BY r.taken_at DESC", (aid,)
    ).fetchall()
    details = db.execute(
        "SELECT d.*, u.name, u.designation FROM answer_details d "
        "LEFT JOIN users u ON u.emp_id = d.emp_id "
        "WHERE d.assessment_id=? ORDER BY d.result_id, d.id", (aid,)
    ).fetchall()
    return a, attempts, details


@app.route("/api/admin/results-report.csv")
@admin_required
def api_admin_results_report_csv():
    """Download the full detailed report as CSV (per-question rows)."""
    aid = request.args.get("id")
    a, attempts, details = _gather_report_rows(aid)
    title = a["title"] if a else "assessment"

    # map result_id -> attempt summary
    amap = {r["id"]: r for r in attempts}

    out = io.StringIO()
    w = csv.writer(out)
    # Header row with items 1-15
    w.writerow([
        "Name", "Emp ID", "Designation", "Assessment", "Date & Time", "Time Taken",
        "Score", "Total", "Percent", "Result",
        "Q#", "Question", "Their Answer", "Correct Answer", "Right/Wrong", "Category"
    ])
    # group details by attempt, number questions within each attempt
    from itertools import groupby
    for rid, group in groupby(details, key=lambda x: x["result_id"]):
        att = amap.get(rid)
        qn = 0
        for drow in group:
            qn += 1
            if att:
                nm, eid, desg = att["name"], att["emp_id"], att["designation"]
                dt = _fmt_dt(att["taken_at"]); tt = _fmt_mins(att["time_taken"])
                sc, tot, pct = att["score"], att["total"], att["percent"]
                res = "PASS" if att["passed"] else "FAIL"
            else:
                nm = eid = desg = dt = tt = ""; sc = tot = pct = ""; res = ""
            w.writerow([
                nm, eid, desg, title, dt, tt, sc, tot, pct, res,
                qn, drow["question_text"], drow["chosen"], drow["correct"],
                "Right" if drow["is_correct"] else "Wrong", drow["category"]
            ])

    from flask import Response
    safe = re.sub(r"[^A-Za-z0-9]+", "_", title).strip("_")
    return Response(
        out.getvalue(), mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="report_{safe}.csv"'}
    )


@app.route("/api/admin/results-report.xlsx")
@admin_required
def api_admin_results_report_xlsx():
    """Download the full detailed report as an Excel file with summary sheets."""
    aid = request.args.get("id")
    a, attempts, details = _gather_report_rows(aid)
    title = a["title"] if a else "assessment"
    amap = {r["id"]: r for r in attempts}

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except Exception:
        return jsonify(ok=False, msg="Excel export not available on server."), 500

    wb = Workbook()

    # ---- Sheet 1: Detailed (per-question) ----
    ws = wb.active
    ws.title = "Detailed Results"
    headers = ["Name", "Emp ID", "Designation", "Assessment", "Date & Time", "Time Taken",
               "Score", "Total", "Percent", "Result",
               "Q#", "Question", "Their Answer", "Correct Answer", "Right/Wrong", "Category"]
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="00AEEF")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = hdr_fill
        c.alignment = Alignment(vertical="center")

    from itertools import groupby
    for rid, group in groupby(details, key=lambda x: x["result_id"]):
        att = amap.get(rid)
        qn = 0
        for drow in group:
            qn += 1
            if att:
                row = [att["name"], att["emp_id"], att["designation"], title,
                       _fmt_dt(att["taken_at"]), _fmt_mins(att["time_taken"]),
                       att["score"], att["total"], att["percent"],
                       "PASS" if att["passed"] else "FAIL"]
            else:
                row = ["", "", "", title, "", "", "", "", "", ""]
            row += [qn, drow["question_text"], drow["chosen"], drow["correct"],
                    "Right" if drow["is_correct"] else "Wrong", drow["category"]]
            ws.append(row)
            # colour the Right/Wrong cell
            rw_cell = ws.cell(row=ws.max_row, column=15)
            if drow["is_correct"]:
                rw_cell.fill = PatternFill("solid", fgColor="D6F3E4")
            else:
                rw_cell.fill = PatternFill("solid", fgColor="FBD9D9")

    widths = [18, 10, 14, 22, 20, 11, 7, 7, 8, 8, 5, 55, 26, 26, 11, 14]
    for i, wdt in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = wdt

    # ---- Sheet 2: Attempts summary (one row per attempt) ----
    ws2 = wb.create_sheet("Attempts Summary")
    ws2.append(["Name", "Emp ID", "Designation", "Date & Time", "Time Taken",
                "Score", "Total", "Percent", "Result"])
    for c in ws2[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = hdr_fill
    for att in attempts:
        ws2.append([att["name"], att["emp_id"], att["designation"],
                    _fmt_dt(att["taken_at"]), _fmt_mins(att["time_taken"]),
                    att["score"], att["total"], att["percent"],
                    "PASS" if att["passed"] else "FAIL"])
    for i, wdt in enumerate([18, 10, 14, 20, 11, 7, 7, 8, 8], start=1):
        ws2.column_dimensions[chr(64 + i)].width = wdt

    # ---- Sheet 3: Hard Questions (#16 — which questions most got wrong) ----
    ws3 = wb.create_sheet("Hard Questions")
    ws3.append(["Question", "Times Answered", "Times Wrong", "Wrong %"])
    for c in ws3[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = hdr_fill
    qstats = {}
    for d in details:
        key = d["question_text"] or f"Q{d['question_id']}"
        st = qstats.setdefault(key, {"n": 0, "wrong": 0})
        st["n"] += 1
        if not d["is_correct"]:
            st["wrong"] += 1
    hard = sorted(qstats.items(), key=lambda kv: (kv[1]["wrong"] / kv[1]["n"]) if kv[1]["n"] else 0, reverse=True)
    for q, st in hard:
        pct = round((st["wrong"] / st["n"]) * 100) if st["n"] else 0
        ws3.append([q, st["n"], st["wrong"], f"{pct}%"])
    for i, wdt in enumerate([60, 15, 12, 10], start=1):
        ws3.column_dimensions[chr(64 + i)].width = wdt

    # ---- Sheet 4: Weak Areas by person (#17 — weak categories per person) ----
    ws4 = wb.create_sheet("Weak Areas")
    ws4.append(["Name", "Emp ID", "Category", "Answered", "Wrong", "Wrong %"])
    for c in ws4[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = hdr_fill
    cat_stats = {}
    for d in details:
        cat = d["category"] or "(uncategorised)"
        key = (d["name"] or d["emp_id"], d["emp_id"], cat)
        st = cat_stats.setdefault(key, {"n": 0, "wrong": 0})
        st["n"] += 1
        if not d["is_correct"]:
            st["wrong"] += 1
    for (nm, eid, cat), st in sorted(cat_stats.items()):
        if st["wrong"] == 0:
            continue  # only show categories where they missed something
        pct = round((st["wrong"] / st["n"]) * 100) if st["n"] else 0
        ws4.append([nm, eid, cat, st["n"], st["wrong"], f"{pct}%"])
    for i, wdt in enumerate([18, 10, 20, 10, 8, 10], start=1):
        ws4.column_dimensions[chr(64 + i)].width = wdt

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    from flask import Response
    safe = re.sub(r"[^A-Za-z0-9]+", "_", title).strip("_")
    return Response(
        bio.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="report_{safe}.xlsx"'}
    )


# ---- Learner side ----
@app.route("/api/my-pending-counts")
@login_required
def api_my_pending_counts():
    """How many induction modules, training modules and assessments the
    logged-in learner still has pending. Powers the red tab badges.
    Module pending = no completion record. Assessment pending = not passed."""
    u = current_user()
    db = get_db()

    def _modules_pending(kind):
        rows = db.execute(
            "SELECT id, roles FROM content_modules WHERE kind=? AND status='live'", (kind,)
        ).fetchall()
        mine = [r for r in rows if _roles_match_designation(r["roles"], u["designation"])]
        if not mine:
            return 0
        done = {c["module_id"] for c in db.execute(
            "SELECT module_id FROM module_completions WHERE emp_id=?", (u["emp_id"],)
        ).fetchall()}
        return sum(1 for r in mine if r["id"] not in done)

    # assessments: matched by role, pending until passed
    arows = db.execute("SELECT id, roles FROM assessments WHERE active=1").fetchall()
    mine_a = [r for r in arows if _roles_match_designation(r["roles"], u["designation"])]
    passed = {p["assessment_id"] for p in db.execute(
        "SELECT DISTINCT assessment_id FROM assessment_results WHERE emp_id=? AND passed=1",
        (u["emp_id"],)
    ).fetchall()}
    assess_pending = sum(1 for r in mine_a if r["id"] not in passed)

    # videos: matched by role, pending until marked watched
    vrows = db.execute("SELECT id, roles FROM videos WHERE status='live'").fetchall()
    mine_v = [r for r in vrows if _roles_match_designation(r["roles"], u["designation"])]
    watched = {w["video_id"] for w in db.execute(
        "SELECT video_id FROM video_completions WHERE emp_id=?", (u["emp_id"],)).fetchall()}
    video_pending = sum(1 for r in mine_v if r["id"] not in watched)

    return jsonify(ok=True,
                   induction=_modules_pending("induction"),
                   training=_modules_pending("training"),
                   assess=assess_pending,
                   videos=video_pending)


@app.route("/api/my-certificates")
@login_required
def api_my_certificates():
    """All certificates this learner earned: assessment passes + completion tracks."""
    u = current_user()
    db = get_db()
    certs = []

    # 1) assessment certificates (passed assessments)
    rows = db.execute(
        "SELECT a.title, MAX(r.percent) AS best, MAX(r.taken_at) AS last_date "
        "FROM assessment_results r JOIN assessments a ON a.id = r.assessment_id "
        "WHERE r.emp_id = ? AND r.passed = 1 "
        "GROUP BY r.assessment_id, a.title ORDER BY last_date DESC",
        (u["emp_id"],)
    ).fetchall()
    for r in rows:
        d = r["last_date"]
        try:
            dt = datetime.fromisoformat(d.replace("Z", "")) if d else None
            date_str = dt.strftime("%d %B %Y") if dt else ""
        except Exception:
            date_str = ""
        certs.append({
            "type": "assessment", "assessment": r["title"], "score": r["best"],
            "date": date_str, "name": u["name"], "emp_id": u["emp_id"]
        })

    # 2) completion-track certificates (issued)
    trows = db.execute(
        "SELECT cert_name, issued_at FROM issued_certificates WHERE emp_id=? ORDER BY issued_at DESC",
        (u["emp_id"],)
    ).fetchall()
    for t in trows:
        d = t["issued_at"]
        try:
            dt = datetime.fromisoformat(d.replace("Z", "")) if d else None
            date_str = dt.strftime("%d %B %Y") if dt else ""
        except Exception:
            date_str = ""
        certs.append({
            "type": "completion", "assessment": t["cert_name"], "score": None,
            "date": date_str, "name": u["name"], "emp_id": u["emp_id"]
        })

    return jsonify(ok=True, certificates=certs)


@app.route("/api/my-assessments")
@login_required
def api_my_assessments():
    """Assessments available to the logged-in learner, with their best result."""
    u = current_user()
    db = get_db()
    rows = db.execute("SELECT * FROM assessments WHERE active=1 ORDER BY created_at DESC").fetchall()
    out = []
    for a in rows:
        if not _assessment_allowed_for(a, u):
            continue
        best = db.execute(
            "SELECT MAX(percent) p, MAX(passed) passed FROM assessment_results "
            "WHERE assessment_id=? AND emp_id=?", (a["id"], u["emp_id"])
        ).fetchone()
        out.append({
            "id": a["id"], "title": a["title"], "description": a["description"],
            "num_questions": a["num_questions"], "pass_percent": a["pass_percent"],
            "time_limit": a["time_limit"],
            "best": best["p"] if best and best["p"] is not None else None,
            "passed": bool(best["passed"]) if best and best["passed"] else False
        })
    return jsonify(ok=True, assessments=out)


@app.route("/api/start-assessment")
@login_required
def api_start_assessment():
    """Return a random, shuffled subset of questions for this learner.
    Correct answers are NOT sent to the browser — scoring happens server-side."""
    u = current_user()
    aid = request.args.get("id")
    db = get_db()
    a = db.execute("SELECT * FROM assessments WHERE id=? AND active=1", (aid,)).fetchone()
    if not a:
        return jsonify(ok=False, msg="Assessment not available."), 404
    if not _assessment_allowed_for(a, u):
        return jsonify(ok=False, msg="This assessment is not assigned to your role."), 403

    qs = db.execute("SELECT * FROM questions WHERE assessment_id=?", (aid,)).fetchall()
    qs = list(qs)
    _random.shuffle(qs)
    take = qs[: a["num_questions"]]

    out = []
    for q in take:
        # build options list and shuffle, keeping track of which is correct
        opts = [("A", q["opt_a"]), ("B", q["opt_b"])]
        if q["opt_c"]:
            opts.append(("C", q["opt_c"]))
        if q["opt_d"]:
            opts.append(("D", q["opt_d"]))
        _random.shuffle(opts)
        out.append({
            "qid": q["id"],
            "question": q["question"],
            # send shuffled options with NEW display letters, hide original correct
            "options": [{"key": chr(65 + idx), "text": text, "_orig": orig}
                        for idx, (orig, text) in enumerate(opts)]
        })

    return jsonify(ok=True, assessment={
        "id": a["id"], "title": a["title"], "time_limit": a["time_limit"],
        "pass_percent": a["pass_percent"], "total": len(out)
    }, questions=out)


@app.route("/api/submit-assessment", methods=["POST"])
@login_required
def api_submit_assessment():
    """Receive answers, score server-side, save result + per-question details, return pass/fail + cert data."""
    u = current_user()
    d = request.get_json(force=True)
    aid = d.get("assessment_id")
    answers = d.get("answers") or {}   # { qid: chosen_orig_letter }
    try:
        time_taken = int(d.get("time_taken") or 0)   # seconds spent, sent by browser
    except (ValueError, TypeError):
        time_taken = 0

    db = get_db()
    a = db.execute("SELECT * FROM assessments WHERE id=?", (aid,)).fetchone()
    if not a:
        return jsonify(ok=False, msg="Assessment not found."), 404

    qids = [int(k) for k in answers.keys()] if answers else []
    score = 0
    total = len(answers)
    qinfo = {}   # qid -> full question row (for saving details)
    if qids:
        placeholders = ",".join("?" * len(qids))
        qrows = db.execute(
            f"SELECT id, question, opt_a, opt_b, opt_c, opt_d, correct, category "
            f"FROM questions WHERE id IN ({placeholders})", qids
        ).fetchall()
        qinfo = {r["id"]: r for r in qrows}
        for qid_str, chosen in answers.items():
            qid = int(qid_str)
            r = qinfo.get(qid)
            if r and str(chosen).upper() == r["correct"]:
                score += 1

    percent = round((score / total) * 100) if total else 0
    passed = 1 if percent >= a["pass_percent"] else 0
    now_iso = datetime.utcnow().isoformat()

    # Save the summary result and get its id back (for linking the details)
    cur = db.execute(
        "INSERT INTO assessment_results (assessment_id,emp_id,score,total,percent,passed,taken_at,time_taken) "
        "VALUES (?,?,?,?,?,?,?,?) RETURNING id",
        (aid, u["emp_id"], score, total, percent, passed, now_iso, time_taken)
    )
    result_id = cur.fetchone()["id"]

    # Save each question's detail (what they chose, correct answer, right/wrong)
    def _letter_text(row, letter):
        m = {"A": row["opt_a"], "B": row["opt_b"], "C": row["opt_c"], "D": row["opt_d"]}
        return m.get((letter or "").upper(), "")

    for qid_str, chosen in answers.items():
        qid = int(qid_str)
        r = qinfo.get(qid)
        if not r:
            continue
        chosen_u = str(chosen).upper()
        is_correct = 1 if chosen_u == r["correct"] else 0
        chosen_full = f"{chosen_u}. {_letter_text(r, chosen_u)}" if chosen_u else "(no answer)"
        correct_full = f"{r['correct']}. {_letter_text(r, r['correct'])}"
        db.execute(
            "INSERT INTO answer_details (result_id,assessment_id,emp_id,question_id,question_text,chosen,correct,is_correct,category,taken_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?)",
            (result_id, aid, u["emp_id"], qid, r["question"], chosen_full, correct_full,
             is_correct, r["category"] or "", now_iso)
        )
    db.commit()

    # passing an assessment may complete a certificate track
    new_certs = _check_and_issue_tracks(u["emp_id"]) if passed else []

    return jsonify(ok=True, score=score, total=total, percent=percent,
                   passed=bool(passed), pass_percent=a["pass_percent"],
                   new_certificates=new_certs,
                   cert={
                       "name": u["name"], "emp_id": u["emp_id"],
                       "assessment": a["title"], "score": percent,
                       "date": datetime.utcnow().strftime("%d %B %Y")
                   } if passed else None)


# ===============================================================
#  NEW — DELIVERY 2: Content Management (modules + videos)
# ===============================================================

def _can_manage_content():
    """Admin OR instructor may add content. Returns the user row or None."""
    u = current_user()
    if u is None:
        return None
    if u["role"] in ("admin", "instructor"):
        return u
    return None


def _content_visible_for(item, u):
    roles = (item["roles"] or "all").strip().lower()
    if roles in ("", "all"):
        return True
    desg = (u["designation"] or "").lower()
    allowed = [r.strip().lower() for r in roles.split(",")]
    return any(a and a in desg for a in allowed)


# ---------- ADMIN/INSTRUCTOR: manage modules ----------
@app.route("/api/admin/modules")
@login_required
def api_admin_modules():
    """List all modules (admin/instructor view). Shows live + pending."""
    u = _can_manage_content()
    if u is None:
        return jsonify(ok=False, msg="Not allowed."), 403
    db = get_db()
    rows = db.execute("SELECT * FROM content_modules ORDER BY kind, sort_order, id").fetchall()
    vids = db.execute("SELECT * FROM videos ORDER BY sort_order, id").fetchall()
    return jsonify(ok=True, modules=[dict(r) for r in rows], videos=[dict(v) for v in vids],
                   role_choices=get_role_choices(), is_admin=(u["role"] == "admin"))


@app.route("/api/admin/upload-file", methods=["POST"])
@login_required
def api_admin_upload_file():
    """Receive a file from the admin/instructor and upload it to Supabase
    Storage. Returns the public URL, which the caller saves as the module link."""
    u = _can_manage_content()
    if u is None:
        return jsonify(ok=False, msg="Not allowed."), 403

    if "file" not in request.files:
        return jsonify(ok=False, msg="No file was selected."), 400
    f = request.files["file"]
    if not f or not f.filename:
        return jsonify(ok=False, msg="No file was selected."), 400

    # validate extension
    import os as _os
    name = f.filename
    ext = _os.path.splitext(name)[1].lower()
    if ext not in ALLOWED_UPLOAD_EXT:
        return jsonify(ok=False, msg=f"File type {ext} is not allowed."), 400

    data = f.read()
    size_mb = len(data) / (1024 * 1024)
    if size_mb > MAX_UPLOAD_MB:
        return jsonify(ok=False, msg=f"File is too large ({size_mb:.1f} MB). Max is {MAX_UPLOAD_MB} MB."), 400

    if not SUPABASE_SECRET or not SUPABASE_URL:
        return jsonify(ok=False, msg="File storage is not configured. Please contact the administrator."), 500

    # Build a safe, unique object path inside the bucket
    import re as _re
    from datetime import datetime as _dt
    safe = _re.sub(r"[^A-Za-z0-9._-]", "_", name)
    stamp = _dt.utcnow().strftime("%Y%m%d%H%M%S")
    object_path = f"{stamp}_{safe}"

    # content type guess
    ctype_map = {
        ".pdf": "application/pdf",
        ".doc": "application/msword",
        ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ".ppt": "application/vnd.ms-powerpoint",
        ".pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        ".xls": "application/vnd.ms-excel",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
        ".gif": "image/gif", ".webp": "image/webp", ".txt": "text/plain",
    }
    ctype = ctype_map.get(ext, "application/octet-stream")

    # Upload to Supabase Storage via its REST endpoint
    import urllib.request
    import urllib.error
    upload_url = f"{SUPABASE_URL}/storage/v1/object/{SUPABASE_BUCKET}/{object_path}"
    req = urllib.request.Request(upload_url, data=data, method="POST")
    req.add_header("Authorization", f"Bearer {SUPABASE_SECRET}")
    req.add_header("apikey", SUPABASE_SECRET)
    req.add_header("Content-Type", ctype)
    req.add_header("x-upsert", "true")
    try:
        urllib.request.urlopen(req, timeout=60)
    except urllib.error.HTTPError as e:
        body = ""
        try:
            body = e.read().decode()[:200]
        except Exception:
            pass
        return jsonify(ok=False, msg=f"Upload failed ({e.code}). {body}"), 502
    except Exception as e:
        return jsonify(ok=False, msg=f"Upload error: {str(e)[:200]}"), 502

    # Public URL for the uploaded file
    public_url = f"{SUPABASE_URL}/storage/v1/object/public/{SUPABASE_BUCKET}/{object_path}"
    # a friendly label for the file type
    label = "pdf" if ext == ".pdf" else ext.lstrip(".")
    return jsonify(ok=True, url=public_url, filename=name, file_type=label,
                   msg="File uploaded.")


@app.route("/api/admin/save-module", methods=["POST"])
@login_required
def api_admin_save_module():
    """Create or update a module. Instructor uploads are 'pending' until admin approves."""
    u = _can_manage_content()
    if u is None:
        return jsonify(ok=False, msg="Not allowed."), 403
    d = request.get_json(force=True)
    mid = d.get("id")
    title = (d.get("title") or "").strip()
    desc = (d.get("description") or "").strip()
    link = (d.get("link") or "").strip()
    kind = (d.get("kind") or "induction").strip()
    file_type = (d.get("file_type") or "").strip()
    roles = (d.get("roles") or "all").strip() or "all"
    try:
        mins = int(d.get("min_minutes") or 0)
    except (ValueError, TypeError):
        mins = 0
    try:
        sort_order = int(d.get("sort_order") or 0)
    except (ValueError, TypeError):
        sort_order = 0

    if not title or not link:
        return jsonify(ok=False, msg="Title and link are required."), 400
    if kind not in ("induction", "training"):
        kind = "induction"

    # Admin content goes live; instructor content is pending approval
    status = "live" if u["role"] == "admin" else "pending"

    db = get_db()
    if mid:
        existing = db.execute("SELECT * FROM content_modules WHERE id=?", (mid,)).fetchone()
        if not existing:
            return jsonify(ok=False, msg="Module not found."), 404
        # if instructor edits, it goes back to pending; admin edits stay live
        new_status = "live" if u["role"] == "admin" else "pending"
        db.execute(
            "UPDATE content_modules SET kind=?,title=?,description=?,link=?,file_type=?,min_minutes=?,roles=?,sort_order=?,status=? WHERE id=?",
            (kind, title, desc, link, file_type, mins, roles, sort_order, new_status, mid)
        )
        db.commit()
        return jsonify(ok=True, msg="Module updated." + ("" if u["role"] == "admin" else " Pending admin approval."))
    else:
        db.execute(
            "INSERT INTO content_modules (kind,title,description,link,file_type,min_minutes,roles,sort_order,status,created_by,created_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            (kind, title, desc, link, file_type, mins, roles, sort_order, status, u["emp_id"], datetime.utcnow().isoformat())
        )
        db.commit()
        msg = "Module added and live." if status == "live" else "Module submitted — pending admin approval."
        return jsonify(ok=True, msg=msg)


@app.route("/api/admin/approve-module", methods=["POST"])
@admin_required
def api_admin_approve_module():
    d = request.get_json(force=True)
    mid = d.get("id")
    db = get_db()
    db.execute("UPDATE content_modules SET status='live' WHERE id=?", (mid,))
    db.commit()
    return jsonify(ok=True, msg="Module approved and live.")


@app.route("/api/admin/delete-module", methods=["POST"])
@login_required
def api_admin_delete_module():
    u = _can_manage_content()
    if u is None:
        return jsonify(ok=False, msg="Not allowed."), 403
    d = request.get_json(force=True)
    mid = d.get("id")
    db = get_db()
    db.execute("DELETE FROM content_modules WHERE id=?", (mid,))
    db.execute("DELETE FROM module_completions WHERE module_id=?", (mid,))
    db.commit()
    return jsonify(ok=True, msg="Module deleted.")


# ---------- ADMIN/INSTRUCTOR: manage videos ----------
@app.route("/api/admin/save-video", methods=["POST"])
@login_required
def api_admin_save_video():
    u = _can_manage_content()
    if u is None:
        return jsonify(ok=False, msg="Not allowed."), 403
    d = request.get_json(force=True)
    vid = d.get("id")
    title = (d.get("title") or "").strip()
    desc = (d.get("description") or "").strip()
    link = (d.get("link") or "").strip()
    roles = (d.get("roles") or "all").strip() or "all"
    try:
        sort_order = int(d.get("sort_order") or 0)
    except (ValueError, TypeError):
        sort_order = 0
    if not title or not link:
        return jsonify(ok=False, msg="Title and link are required."), 400

    status = "live" if u["role"] == "admin" else "pending"
    db = get_db()
    if vid:
        new_status = "live" if u["role"] == "admin" else "pending"
        db.execute("UPDATE videos SET title=?,description=?,link=?,roles=?,sort_order=?,status=? WHERE id=?",
                   (title, desc, link, roles, sort_order, new_status, vid))
    else:
        db.execute("INSERT INTO videos (title,description,link,roles,sort_order,status,created_by,created_at) VALUES (?,?,?,?,?,?,?,?)",
                   (title, desc, link, roles, sort_order, status, u["emp_id"], datetime.utcnow().isoformat()))
    db.commit()
    return jsonify(ok=True, msg="Video saved." + ("" if status == "live" or vid else " Pending admin approval."))


@app.route("/api/admin/approve-video", methods=["POST"])
@admin_required
def api_admin_approve_video():
    d = request.get_json(force=True)
    vid = d.get("id")
    db = get_db()
    db.execute("UPDATE videos SET status='live' WHERE id=?", (vid,))
    db.commit()
    return jsonify(ok=True)


@app.route("/api/admin/delete-video", methods=["POST"])
@login_required
def api_admin_delete_video():
    u = _can_manage_content()
    if u is None:
        return jsonify(ok=False, msg="Not allowed."), 403
    d = request.get_json(force=True)
    vid = d.get("id")
    db = get_db()
    db.execute("DELETE FROM videos WHERE id=?", (vid,))
    db.commit()
    return jsonify(ok=True)


# ---------- LEARNER: view modules + videos ----------
@app.route("/api/content/<kind>")
@login_required
def api_content(kind):
    """Learner-facing modules of a kind ('induction' or 'training'), with completion state."""
    u = current_user()
    if kind not in ("induction", "training"):
        return jsonify(ok=False, msg="Unknown content type."), 400
    db = get_db()
    rows = db.execute(
        "SELECT * FROM content_modules WHERE kind=? AND status='live' ORDER BY sort_order, id", (kind,)
    ).fetchall()
    done = {r["module_id"] for r in db.execute(
        "SELECT module_id FROM module_completions WHERE emp_id=?", (u["emp_id"],)).fetchall()}
    out = []
    for m in rows:
        if not _content_visible_for(m, u):
            continue
        d = dict(m); d["completed"] = m["id"] in done
        out.append(d)
    return jsonify(ok=True, modules=out)


@app.route("/api/content/videos")
@login_required
def api_content_videos():
    u = current_user()
    db = get_db()
    rows = db.execute("SELECT * FROM videos WHERE status='live' ORDER BY sort_order, id").fetchall()
    done = {r["video_id"] for r in db.execute(
        "SELECT video_id FROM video_completions WHERE emp_id=?", (u["emp_id"],)).fetchall()}
    out = []
    for v in rows:
        if not _content_visible_for(v, u):
            continue
        d = dict(v); d["completed"] = v["id"] in done
        out.append(d)
    return jsonify(ok=True, videos=out)


@app.route("/api/content/complete-video", methods=["POST"])
@login_required
def api_content_complete_video():
    """Mark a video watched for this learner (the 'Mark as watched' button)."""
    u = current_user()
    d = request.get_json(force=True)
    vid = d.get("video_id")
    if not vid:
        return jsonify(ok=False, msg="Missing video."), 400
    db = get_db()
    already = db.execute(
        "SELECT id FROM video_completions WHERE video_id=? AND emp_id=?", (vid, u["emp_id"])
    ).fetchone()
    if not already:
        db.execute(
            "INSERT INTO video_completions (video_id, emp_id, completed_at) VALUES (?,?,?)",
            (vid, u["emp_id"], datetime.utcnow().isoformat())
        )
        db.commit()
    return jsonify(ok=True)


@app.route("/api/content/complete", methods=["POST"])
@login_required
def api_content_complete():
    """Mark a module complete for this learner (called after the timer elapses)."""
    u = current_user()
    d = request.get_json(force=True)
    mid = d.get("module_id")
    if not mid:
        return jsonify(ok=False, msg="Missing module."), 400
    db = get_db()
    already = db.execute("SELECT 1 FROM module_completions WHERE module_id=? AND emp_id=?",
                         (mid, u["emp_id"])).fetchone()
    if not already:
        db.execute("INSERT INTO module_completions (module_id,emp_id,completed_at) VALUES (?,?,?)",
                   (mid, u["emp_id"], datetime.utcnow().isoformat()))
        db.commit()
    # check if this completion earned any track certificate
    newly = _check_and_issue_tracks(u["emp_id"])
    return jsonify(ok=True, msg="Marked complete.", new_certificates=newly)


# ===============================================================
#  Certificate Tracks (completion certificates) — Delivery
# ===============================================================

def _check_and_issue_tracks(emp_id):
    """Check all live tracks for this user; issue any newly-earned certificates.
    Returns list of newly issued cert names."""
    db = get_db()
    u = db.execute("SELECT * FROM users WHERE emp_id=?", (emp_id,)).fetchone()
    if not u:
        return []
    tracks = db.execute("SELECT * FROM certificate_tracks WHERE status='live'").fetchall()
    newly = []
    for t in tracks:
        # already issued?
        got = db.execute("SELECT 1 FROM issued_certificates WHERE track_id=? AND emp_id=?",
                         (t["id"], emp_id)).fetchone()
        if got:
            continue
        # does this track apply to the user's role?
        roles = (t["roles"] or "all").strip().lower()
        if roles not in ("", "all"):
            desg = (u["designation"] or "").lower()
            allowed = [r.strip().lower() for r in roles.split(",")]
            if not any(a and a in desg for a in allowed):
                continue

        qualifies = True

        # requirement 1: complete all matching live modules of this kind+roles
        if t["require_modules"]:
            mods = db.execute("SELECT * FROM content_modules WHERE kind=? AND status='live'",
                              (t["kind"],)).fetchall()
            # filter modules that apply to this user's role
            relevant = []
            for m in mods:
                mroles = (m["roles"] or "all").strip().lower()
                if mroles in ("", "all"):
                    relevant.append(m)
                else:
                    desg = (u["designation"] or "").lower()
                    if any(a.strip() and a.strip() in desg for a in mroles.split(",")):
                        relevant.append(m)
            if len(relevant) == 0:
                qualifies = False  # nothing to complete yet
            else:
                done_ids = {r["module_id"] for r in db.execute(
                    "SELECT module_id FROM module_completions WHERE emp_id=?", (emp_id,)).fetchall()}
                for m in relevant:
                    if m["id"] not in done_ids:
                        qualifies = False
                        break

        # requirement 2: pass a specific assessment (if set)
        if qualifies and t["require_assessment_id"]:
            passed = db.execute(
                "SELECT 1 FROM assessment_results WHERE emp_id=? AND assessment_id=? AND passed=1 LIMIT 1",
                (emp_id, t["require_assessment_id"])).fetchone()
            if not passed:
                qualifies = False

        if qualifies:
            db.execute("INSERT INTO issued_certificates (track_id,emp_id,cert_name,issued_at) VALUES (?,?,?,?)",
                       (t["id"], emp_id, t["cert_name"], datetime.utcnow().isoformat()))
            newly.append(t["cert_name"])

    if newly:
        db.commit()
    return newly


@app.route("/api/admin/cert-tracks")
@login_required
def api_admin_cert_tracks():
    """List certificate tracks (admin/instructor)."""
    u = _can_manage_content()
    if u is None:
        return jsonify(ok=False, msg="Not allowed."), 403
    db = get_db()
    rows = db.execute("SELECT * FROM certificate_tracks ORDER BY created_at DESC").fetchall()
    # attach assessment titles + issue counts
    out = []
    for t in rows:
        d = dict(t)
        if t["require_assessment_id"]:
            a = db.execute("SELECT title FROM assessments WHERE id=?", (t["require_assessment_id"],)).fetchone()
            d["assessment_title"] = a["title"] if a else "(deleted)"
        else:
            d["assessment_title"] = None
        d["issued_count"] = db.execute("SELECT COUNT(*) c FROM issued_certificates WHERE track_id=?", (t["id"],)).fetchone()["c"]
        out.append(d)
    # also send assessments list for the dropdown
    assessments = db.execute("SELECT id,title FROM assessments ORDER BY title").fetchall()
    return jsonify(ok=True, tracks=out, assessments=[dict(a) for a in assessments],
                   role_choices=get_role_choices(), is_admin=(u["role"] == "admin"))


@app.route("/api/admin/save-cert-track", methods=["POST"])
@login_required
def api_admin_save_cert_track():
    u = _can_manage_content()
    if u is None:
        return jsonify(ok=False, msg="Not allowed."), 403
    d = request.get_json(force=True)
    tid = d.get("id")
    cert_name = (d.get("cert_name") or "").strip()
    kind = (d.get("kind") or "training").strip()
    roles = (d.get("roles") or "all").strip() or "all"
    require_modules = 1 if d.get("require_modules", True) else 0
    req_assess = d.get("require_assessment_id")
    if req_assess in ("", "none", "0", 0):
        req_assess = None

    if not cert_name:
        return jsonify(ok=False, msg="Certificate name is required."), 400
    if kind not in ("induction", "training"):
        kind = "training"
    if not require_modules and not req_assess:
        return jsonify(ok=False, msg="Pick at least one requirement (modules and/or an assessment)."), 400

    status = "live" if u["role"] == "admin" else "pending"
    db = get_db()
    if tid:
        new_status = "live" if u["role"] == "admin" else "pending"
        db.execute("UPDATE certificate_tracks SET cert_name=?,kind=?,roles=?,require_modules=?,require_assessment_id=?,status=? WHERE id=?",
                   (cert_name, kind, roles, require_modules, req_assess, new_status, tid))
    else:
        db.execute("INSERT INTO certificate_tracks (cert_name,kind,roles,require_modules,require_assessment_id,status,created_by,created_at) VALUES (?,?,?,?,?,?,?,?)",
                   (cert_name, kind, roles, require_modules, req_assess, status, u["emp_id"], datetime.utcnow().isoformat()))
    db.commit()
    msg = "Certificate track saved." if status == "live" else "Track submitted — pending admin approval."
    return jsonify(ok=True, msg=msg)


@app.route("/api/admin/approve-cert-track", methods=["POST"])
@admin_required
def api_admin_approve_cert_track():
    d = request.get_json(force=True)
    db = get_db()
    db.execute("UPDATE certificate_tracks SET status='live' WHERE id=?", (d.get("id"),))
    db.commit()
    return jsonify(ok=True, msg="Track approved.")


@app.route("/api/admin/delete-cert-track", methods=["POST"])
@login_required
def api_admin_delete_cert_track():
    u = _can_manage_content()
    if u is None:
        return jsonify(ok=False, msg="Not allowed."), 403
    d = request.get_json(force=True)
    db = get_db()
    db.execute("DELETE FROM certificate_tracks WHERE id=?", (d.get("id"),))
    db.execute("DELETE FROM issued_certificates WHERE track_id=?", (d.get("id"),))
    db.commit()
    return jsonify(ok=True, msg="Track deleted.")


# ---------------------------------------------------------------
#  Start
# ---------------------------------------------------------------
init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
